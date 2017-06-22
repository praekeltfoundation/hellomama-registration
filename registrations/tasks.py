import json
import requests
import uuid
from datetime import datetime

import pika
from celery.task import Task
from celery.utils.log import get_task_logger
from django.conf import settings
from go_http.metrics import MetricsApiClient
from openpyxl import load_workbook
from io import BytesIO

from hellomama_registration import utils
from .graphite import RetentionScheme
from .models import (Registration, SubscriptionRequest, Source,
                     ThirdPartyRegistrationError)
from .metrics import MetricGenerator, send_metric
from .serializers import RegistrationSerializer

logger = get_task_logger(__name__)


def is_valid_date(date):
    try:
        datetime.strptime(date, "%Y%m%d")
        return True
    except:
        return False


def is_valid_uuid(id):
    return len(id) == 36 and id[14] == '4' and id[19] in ['a', 'b', '8', '9']


def is_valid_lang(lang):
    return lang in settings.LANGUAGES


def is_valid_msg_type(msg_type):
    return msg_type in settings.MSG_TYPES


def is_valid_msg_receiver(msg_receiver):
    return msg_receiver in settings.RECEIVER_TYPES


def is_valid_loss_reason(loss_reason):
    return loss_reason in ['miscarriage', 'stillborn', 'baby_died']


def is_valid_state(state):
    return state in settings.STATES


def is_valid_role(role):
    return role in settings.ROLES


class ValidateRegistration(Task):
    """ Task to validate a registration model entry's registration
    data.
    """
    name = "hellomama_registration.registrations.tasks.validate_registration"

    def check_field_values(self, fields, registration_data):
        failures = []
        for field in fields:
            if field in ["receiver_id", "operator_id"]:
                if not is_valid_uuid(registration_data[field]):
                    failures.append(field)
            if field == "language":
                if not is_valid_lang(registration_data[field]):
                    failures.append(field)
            if field == "msg_type":
                if not is_valid_msg_type(registration_data[field]):
                    failures.append(field)
            if field in ["last_period_date", "baby_dob"]:
                if not is_valid_date(registration_data[field]):
                    failures.append(field)
                elif field == "last_period_date":
                    # Check last_period_date is in valid week range
                    preg_weeks = utils.calc_pregnancy_week_lmp(
                        utils.get_today(), registration_data[field])
                    if not (settings.PREBIRTH_MIN_WEEKS <= preg_weeks <=
                            settings.PREBIRTH_MAX_WEEKS):
                        failures.append("last_period_date out of range")
                elif field == "baby_dob":
                    # Check baby_dob is in valid week range
                    preg_weeks = utils.calc_baby_age(
                        utils.get_today(), registration_data[field])
                    if not (settings.POSTBIRTH_MIN_WEEKS <= preg_weeks <=
                            settings.POSTBIRTH_MAX_WEEKS):
                        failures.append("baby_dob out of range")
            if field == "msg_receiver":
                if not is_valid_msg_receiver(registration_data[field]):
                    failures.append(field)
            if field == "loss_reason":
                if not is_valid_loss_reason(registration_data[field]):
                    failures.append(field)
        return failures

    def validate(self, registration):
        """ Validates that all the required info is provided for a
        registration.
        """
        data_fields = registration.data.keys()
        fields_general = ["receiver_id", "operator_id",
                          "language", "msg_type"]
        fields_prebirth = ["last_period_date", "msg_receiver"]
        fields_postbirth = ["baby_dob", "msg_receiver"]
        fields_loss = ["loss_reason"]

        hw_pre = list(set(fields_general) | set(fields_prebirth))
        hw_post = list(set(fields_general) | set(fields_postbirth))
        pbl_loss = list(set(fields_general) | set(fields_loss))

        # Check if mother_id is a valid UUID
        if not is_valid_uuid(registration.mother_id):
            registration.data["invalid_fields"] = "Invalid UUID mother_id"
            registration.save()
            return False

        if "msg_receiver" in registration.data.keys():
            # Reject registrations on behalf of mother that does not have a
            # unique id for the mother
            if (registration.data["msg_receiver"] in [
                "father_only", "friend_only", "family_only"] and
               registration.mother_id == registration.data["receiver_id"]):
                registration.data["invalid_fields"] = "mother requires own id"
                registration.save()
                return False
            # Reject registrations where the mother is the receiver but the
            # mother_id and receiver_id differs
            elif (registration.data["msg_receiver"] == "mother_only" and
                  registration.mother_id != registration.data["receiver_id"]):
                registration.data["invalid_fields"] = "mother_id should be " \
                    "the same as receiver_id"
                registration.save()
                return False

        # HW registration, prebirth
        if (registration.stage == "prebirth" and
                registration.source.authority in ["hw_limited", "hw_full"] and
                set(hw_pre).issubset(data_fields)):  # ignore extra data

            invalid_fields = self.check_field_values(
                hw_pre, registration.data)
            if invalid_fields == []:
                registration.data["reg_type"] = "hw_pre"
                registration.data["preg_week"] = utils.calc_pregnancy_week_lmp(
                    utils.get_today(), registration.data["last_period_date"])
                registration.validated = True
                registration.save()
                return True
            else:
                registration.data["invalid_fields"] = invalid_fields
                registration.save()
                return False
        # HW registration, postbirth
        elif (registration.stage == "postbirth" and
              registration.source.authority in ["hw_limited", "hw_full"] and
              set(hw_post).issubset(data_fields)):
            invalid_fields = self.check_field_values(
                hw_post, registration.data)
            if invalid_fields == []:
                registration.data["reg_type"] = "hw_post"
                registration.data["baby_age"] = utils.calc_baby_age(
                    utils.get_today(), registration.data["baby_dob"])
                registration.validated = True
                registration.save()
                return True
            else:
                registration.data["invalid_fields"] = invalid_fields
                registration.save()
                return False
        # Loss registration
        elif (registration.stage == "loss" and
              registration.source.authority in ["patient", "advisor"] and
              set(pbl_loss).issubset(data_fields)):
            invalid_fields = self.check_field_values(
                pbl_loss, registration.data)
            if invalid_fields == []:
                registration.data["reg_type"] = "pbl_loss"
                registration.validated = True
                registration.save()
                return True
            else:
                registration.data["invalid_fields"] = invalid_fields
                registration.save()
                return False
        else:
            registration.data[
                "invalid_fields"] = "Invalid combination of fields"
            registration.save()
            return False

    def create_subscriptionrequests(self, registration):
        """ Create SubscriptionRequest(s) based on the
        validated registration.
        """

        voice_days, voice_times = registration.get_voice_days_and_times()
        weeks = registration.get_weeks_pregnant_or_age()

        mother_short_name = utils.get_messageset_short_name(
            registration.stage, 'mother', registration.data["msg_type"],
            weeks, voice_days, voice_times
        )

        mother_msgset_id, mother_msgset_schedule, next_sequence_number =\
            utils.get_messageset_schedule_sequence(mother_short_name, weeks)

        mother_sub = {
            "identity": registration.mother_id,
            "messageset": mother_msgset_id,
            "next_sequence_number": next_sequence_number,
            "lang": registration.data["language"],
            "schedule": mother_msgset_schedule,
            "metadata": {}
        }

        # Add mother welcome message
        if 'voice_days' in registration.data and \
                registration.data["voice_days"] != "":
            mother_sub["metadata"]["prepend_next_delivery"] = \
                "%s/static/audio/registration/%s/welcome_mother.mp3" % (
                settings.PUBLIC_HOST,
                registration.data["language"])
        else:
            # mother_identity =
            if registration.data["msg_receiver"] in [
                    "father_only", "friend_only", "family_only"]:
                to_addr = utils.get_identity_address(
                    registration.data["receiver_id"])
            else:
                to_addr = utils.get_identity_address(
                    registration.mother_id)
            payload = {
                "to_addr": to_addr,
                "content": settings.MOTHER_WELCOME_TEXT_NG_ENG,
                "metadata": {}
            }
            utils.post_message(payload)

        SubscriptionRequest.objects.create(**mother_sub)

        if registration.data["msg_receiver"] != 'mother_only':
            if 'preg_week' in registration.data:
                weeks = registration.data["preg_week"]
            else:
                weeks = registration.data["baby_age"]

            household_short_name = utils.get_messageset_short_name(
                registration.stage, 'household', registration.data["msg_type"],
                registration.data["preg_week"], "fri", "9_11"
            )

            household_msgset_id, household_msgset_schedule, seq_number =\
                utils.get_messageset_schedule_sequence(
                    household_short_name, weeks)
            household_sub = {
                "identity": registration.data["receiver_id"],
                "messageset": household_msgset_id,
                "next_sequence_number": seq_number,
                "lang": registration.data["language"],
                "schedule": household_msgset_schedule,
                "metadata": {}
            }
            # Add household welcome message
            household_sub["metadata"]["prepend_next_delivery"] = \
                "%s/static/audio/registration/%s/welcome_household.mp3" % (
                settings.PUBLIC_HOST,
                registration.data["language"])
            SubscriptionRequest.objects.create(**household_sub)
            return "2 SubscriptionRequests created"

        return "1 SubscriptionRequest created"

    def run(self, registration_id, **kwargs):
        """ Sets the registration's validated field to True if
        validation is successful.
        """
        l = self.get_logger(**kwargs)
        l.info("Looking up the registration")
        registration = Registration.objects.get(id=registration_id)
        reg_validates = self.validate(registration)

        validation_string = "Validation completed - "
        if reg_validates:
            self.create_subscriptionrequests(registration)
            validation_string += "Success"
        else:
            validation_string += "Failure"

        return validation_string

validate_registration = ValidateRegistration()


class DeliverHook(Task):
    def run(self, target, payload, instance_id=None, hook_id=None, **kwargs):
        """
        target:     the url to receive the payload.
        payload:    a python primitive data structure
        instance_id:   a possibly None "trigger" instance ID
        hook_id:       the ID of defining Hook object
        """
        requests.post(
            url=target,
            data=json.dumps(payload),
            headers={
                'Content-Type': 'application/json',
                'Authorization': 'Token %s' % settings.HOOK_AUTH_TOKEN
            }
        )


def deliver_hook_wrapper(target, payload, instance, hook):
    if isinstance(instance.id, uuid.UUID):
        instance_id = str(instance.id)
    else:
        instance_id = instance.id
    kwargs = dict(target=target, payload=payload,
                  instance_id=instance_id, hook_id=hook.id)
    DeliverHook.apply_async(kwargs=kwargs)


def get_metric_client(session=None):
    return MetricsApiClient(
        auth_token=settings.METRICS_AUTH_TOKEN,
        api_url=settings.METRICS_URL,
        session=session)


class FireMetric(Task):

    """ Fires a metric using the MetricsApiClient
    """
    name = "hellomama_registration.tasks.fire_metric"

    def run(self, metric_name, metric_value, session=None, **kwargs):
        metric_value = float(metric_value)
        metric = {
            metric_name: metric_value
        }
        metric_client = get_metric_client(session=session)
        metric_client.fire(metric)
        return "Fired metric <%s> with value <%s>" % (
            metric_name, metric_value)

fire_metric = FireMetric()


class RepopulateMetrics(Task):
    """
    Repopulates historical metrics.
    """
    name = 'registrations.tasks.repopulate_metrics'

    def generate_and_send(
            self, amqp_url, prefix, metric_name, start, end):
        """
        Generates the value for the specified metric, and sends it.
        """
        try:
            value = MetricGenerator().generate_metric(metric_name, start, end)
        except requests.exceptions.RequestException:
            # If we have an issue contacting an external service for this
            # metric, just skip it.
            return
        except ValueError:
            # If an external service doesn't return an valid JSON response for
            # this metric, just skip it.
            return

        timestamp = start + (end - start) / 2
        send_metric(amqp_url, prefix, metric_name, value, timestamp)

    def run(
            self, amqp_url, prefix, metric_names, graphite_retentions,
            **kwargs):
        parameters = pika.URLParameters(amqp_url)
        connection = pika.BlockingConnection(parameters)
        amqp_channel = connection.channel()

        ret = RetentionScheme(graphite_retentions)
        for start, end in ret.get_buckets():
            for metric in metric_names:
                self.generate_and_send(
                    amqp_channel, prefix, metric, start, end)

        connection.close()

repopulate_metrics = RepopulateMetrics()


class PullThirdPartyRegistrations(Task):

    def get_or_create_identity(self, msisdn, details={}):

        if not msisdn:
            return None

        msisdn = utils.normalize_msisdn(msisdn, '234')

        params = {"details__addresses__msisdn": msisdn}
        identities = utils.search_identities(params)

        for identity in identities:
            if details:
                identity['details'].update(details)
                utils.patch_identity(identity['id'], identity['details'])
            return identity['id']

        identity = {
            'details': {
                'default_addr_type': 'msisdn',
                'addresses': {
                    'msisdn': {
                        msisdn: {'default': True}
                    }
                }
            }
        }
        identity['details'].update(details)
        identity = utils.create_identity(identity)['id']
        return identity

    def get_language(self, language):
        return {'english': 'eng_NG',
                'igbo': 'ibo_NG',
                'pidgin': 'pcm_NG'}.get(language, language)

    def get_msg_type(self, msg_type):
        return 'audio' if msg_type == 'voice' else msg_type

    def get_voice_times(self, time):
        return {'9-11am': '9_11',
                '2-5pm': '2_5',
                '6-8pm': '6_8'}.get(time, time)

    def get_voice_days(self, days):
        return {'monday_and_wednesday': 'mon_wed',
                'tuesday_and_thursday': 'tue_thu'}.get(days, days)

    def get_receiver(self, receiver):
        return {
            "mother_and_father": "mother_father",
            "mother_and_family": "mother_family",
            "mother_and_friend": "mother_friend",
        }.get(receiver, receiver)

    def get_data(self):
        url = settings.THIRDPARTY_REGISTRATIONS_URL
        username = settings.THIRDPARTY_REGISTRATIONS_USER
        password = settings.THIRDPARTY_REGISTRATIONS_PASSWORD

        response = requests.get(url, verify=False, auth=(username, password))

        wb = load_workbook(filename=BytesIO(response.content))
        ws = wb.get_sheet_by_name('Forms')

        columns = []
        header_row = ws[1]
        for cell in header_row:
            columns.append(cell.value.replace('form.', ''))

        data = []

        for row in ws.iter_rows(min_row=2):
            reg = {}
            for i, column in enumerate(columns):
                reg[column] = row[i].value

            data.append(reg)

        return data

    def run(self, user_id, **kwargs):
        data = self.get_data()

        source = Source.objects.get(user=user_id)

        loop_error = None

        for line in data:

            try:
                operator_id = self.get_or_create_identity(
                    line['health_worker_phone_number'])

                language = self.get_language(line['preferred_msg_language'])
                receiver = self.get_receiver(line['message_receiver'])

                identity_details = {
                    "default_addr_type": "msisdn",
                    "operator_id": operator_id,
                    "preferred_language": language,
                    "receiver_role": "mother",
                    "gravida": line['gravida'],
                }

                mother_id = self.get_or_create_identity(
                    line['mothers_phone_number'], identity_details)

                msg_type = self.get_msg_type(line['preferred_msg_type'])

                reg_info = {
                    "stage": line['type_of_registration'],
                    "mother_id": mother_id,
                    "source": source.id,
                    "data": {
                        "msg_receiver": receiver,
                        "operator_id": operator_id,
                        "language": language,
                        "msg_type": msg_type
                    }
                }

                if line['gatekeeper_phone_number']:
                    identity_details["receiver_role"] = receiver.replace(
                        'mother_', '')
                    identity_details["linked_to"] = mother_id
                    receiver_id = self.get_or_create_identity(
                        line['gatekeeper_phone_number'],
                        identity_details)
                    reg_info['data']['receiver_id'] = receiver_id

                if msg_type == 'audio':
                    reg_info['data']['voice_times'] = self.get_voice_times(
                        line['message_time'])
                    reg_info['data']['voice_days'] = self.get_voice_days(
                        line['message_days'])

                if line['type_of_registration'] == 'prebirth':
                    reg_info['data']['last_period_date'] = \
                        utils.calc_date_from_pregnancy_week(
                            utils.get_today(),
                            line["pregnancy_week"]).strftime("%Y%m%d")
                else:
                    reg_info['data']['baby_dob'] = utils.calc_baby_dob(
                        utils.get_today(),
                        line["pregnancy_week"]).strftime("%Y%m%d")

                serializer = RegistrationSerializer(data=reg_info)
                serializer.is_valid(raise_exception=True)
                serializer.save()
            except Exception as error:
                loop_error = error
                line['error'] = json.dumps(error.detail)
                ThirdPartyRegistrationError.objects.create(**{'data': line})

        if loop_error:
            raise

pull_third_party_registrations = PullThirdPartyRegistrations()
